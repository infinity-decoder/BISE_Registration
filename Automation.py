"""Python program to demonstrate selenium"""

import openpyxl
import time
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.remote.webelement import WebElement

driver_path = Service("C:\\WebDrivers\\chromedriver.exe") # change the path according to yours.
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(chrome_options=options, service=driver_path)
driver.maximize_window()
actions = ActionChains(driver)
driver.get("http://registration.bisefsd.edu.pk/Account/Login.aspx")


"""controls on website"""
username = driver.find_element("id", "ContentPlaceHolder1_LoginUser_UserName")
username.send_keys("101268")
password = driver.find_element("id", "ContentPlaceHolder1_LoginUser_Password")
password.send_keys("azam.shams")
driver.find_element("name", "ctl00$ContentPlaceHolder1$LoginUser$ctl01").click()
driver.get("http://registration.bisefsd.edu.pk/InstituteForms/RegistrationList.aspx")
driver.get("http://registration.bisefsd.edu.pk/InstituteForms/RegistrationNew.aspx")

# #import data from excel file
path = "F:\\SSC_images\\9th_Registration.xlsx"   # enter the path of Excel file
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# dictionary that contains multiple lists
list_of_rows = {}

for x in range(1,max_row+1):
    list_of_rows['row'+str(x)] = []

# print("Printing lists of rows ",list_of_rows)
for r in range(2,max_row+1):
    for a in range(1, max_col+1):
        cell_obj = sheet_obj.cell(row = r, column = a)
        if a == 8:
            cell_value= "0"+str(cell_obj.value)
            list_of_rows['row' + str(r)].append(cell_value)
        else:
            cell_value = cell_obj.value
            list_of_rows['row' + str(r)].append(cell_value)

##printing dictionary loop
for x in range(1, max_row+1):
    print("Key Name: " , list(list_of_rows.keys())[x-1])
    print(list_of_rows['row'+str(x)])

#Need a list that contains all the ids of elements
element_ids = ["ContentPlaceHolder1_TextBoxName",
               "ContentPlaceHolder1_TextBoxFather",
               "ContentPlaceHolder1_TextBoxBirthDate",
               "ContentPlaceHolder1_imgPhoto",
               "ContentPlaceHolder1_TextBoxFCnic",
               "ContentPlaceHolder1_TextBoxBForm",
               "ContentPlaceHolder1_TextBoxClassRNo",
               "ContentPlaceHolder1_TextBoxMobile",
               "ContentPlaceHolder1_TextBoxAdmissionDate",
               "ContentPlaceHolder1_Radio_Hafiz",
               "ContentPlaceHolder1_TextBoxIdMark",
               "ContentPlaceHolder1_TextBoxAddress",
               "//select[@id='ContentPlaceHolder1_ddlDistrict']/option[text()='TTSINGH']",
               "//select[@id='ContentPlaceHolder1_ddlTehsil']/option[text()='PIR MAHAL']",
               "Dummy string for subject select iteration"
               ]

print("Printing List Length ... ", len(element_ids))
print("last element of list : Subject: ", list_of_rows['row2'][-1])
#Start putting in fields
for keys,values in list_of_rows.items():
    for counter,val in enumerate(values):
        # "ContentPlaceHolder1_TextBoxName",
        if values.index(val) == 0:
            #print("The Zeroth index of "+str(counter)+ str(val))
            driver.find_element("id", element_ids[0]).send_keys(val)
        # "ContentPlaceHolder1_TextBoxFather",
        elif values.index(val) == 1:
            #print("1st index printin..."+str(val))
            driver.find_element("id", element_ids[1]).send_keys(val)
        # "ContentPlaceHolder1_TextBoxBirthDate",
        elif values.index(val) == 2:
            beautiful_date = val.strftime('%d/%m/%Y')
            #print("Date in Beautiful Format...####",beautiful_date)
            driver.implicitly_wait(0.1)
            datefield = driver.find_element("id", element_ids[2])
            ActionChains(driver).move_to_element(datefield).click().send_keys(beautiful_date).perform()
            driver.implicitly_wait(0.1)
        # "ContentPlaceHolder1_imgPhoto",

        #btnModalPopup
        elif values.index(val) == 3:
            time.sleep(1)
            driver.find_element("id", "btnModalPopup").click()
            driver.implicitly_wait(1.0)
            WebElement.chooseFile = driver.find_element("xpath", ".//input[@id='ContentPlaceHolder1_CropPhoto_image_file_pv']")\
                .send_keys(val)
            driver.find_element("id", "btnSelect",).click()

        #FatherCNIC_index=4 , "ContentPlaceHolder1_TextBoxFCnic"
        elif values.index(val) == 4:
            driver.find_element("id", element_ids[4]).send_keys(val)
        #                "ContentPlaceHolder1_TextBoxBForm",
        elif values.index(val) == 5:
            driver.find_element("id", element_ids[5]).send_keys(val)
        #                "ContentPlaceHolder1_TextBoxClassRNo",
        elif values.index(val) == 6:
            driver.find_element("id", element_ids[6]).send_keys(val)
        #                "ContentPlaceHolder1_TextBoxMobile",
        elif values.index(val) == 7:
            driver.find_element("id", element_ids[7]).send_keys(val)
            driver.implicitly_wait(1.0)
        # Admission_Date
        elif values.index(val) == 8:
            beautiful_date = val.strftime('%d/%m/%Y')
            datefield = driver.find_element("id", element_ids[8])
            ActionChains(driver).move_to_element(datefield).click().send_keys(beautiful_date).perform()
            driver.implicitly_wait(0.1)

        # IDMark
        elif values.index(val) == 9:
            driver.find_element("id", element_ids[10]).send_keys(val)

        # Address
        elif values.index(val) == 11:
            driver.find_element("id", element_ids[11]).send_keys(val)
            #Selecting District and Tehsil Here
            driver.implicitly_wait(1.0)
            driver.find_element("xpath", ".//select[@id='ContentPlaceHolder1_ddlDistrict']/option[text()='TTSINGH']").click()
            driver.implicitly_wait(7.0)
            time.sleep(5)
            #Tehsil
            driver.find_element("xpath", ".//select[@id='ContentPlaceHolder1_ddlTehsil']/option[text()='PIR MAHAL']").click()


        # "ContentPlaceHolder1_ddlDistrict"SelectDistrict,

        #Select subject iteration
        elif values.index(val) == 12:
            if val == "Biology" or val == "Computer":
                print("Subject is :", val)
                time.sleep(3)
                if val == "Biology":
                    select = Select(driver.find_element("id", "ContentPlaceHolder1_ddlSubject18"))
                    select.select_by_visible_text('Biology')
                elif val == "Computer":
                    select = Select(driver.find_element("id", "ContentPlaceHolder1_ddlSubject18"))
                    select.select_by_visible_text('Computer Science')
            else:
                select = Select(driver.find_element("id", "ContentPlaceHolder1_ddlGroup"))
                select.select_by_visible_text('GENERAL')
                driver.implicitly_wait(7.0)
                time.sleep(5)

                select_17 = Select(driver.find_element("id", "ContentPlaceHolder1_ddlSubject17"))
                select_17.select_by_visible_text('Islamic Studies')
                select_18 = Select(driver.find_element("id", "ContentPlaceHolder1_ddlSubject18"))
                select_18.select_by_visible_text('Physical Education')

            #Code for Submit button should be written here:
            driver.find_element("id", "ContentPlaceHolder1_ButtonSave").click()
            # wait for refresh
            driver.implicitly_wait(5)
            # Add New registration
            driver.get("http://registration.bisefsd.edu.pk/InstituteForms/RegistrationNew.aspx")
            print("Save Clicked")
            # break the loop on completion







